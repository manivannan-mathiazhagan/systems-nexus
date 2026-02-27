import os
import re
from openpyxl import load_workbook

def substitute(text, label):
    """Substitute &ParameterLabel (case-insensitive) with the actual label."""
    return re.sub(r"&parameterlabel", label, text, flags=re.IGNORECASE) if text else ""

def get_titles_with_substitution(excel_path, table_id):
    wb = load_workbook(excel_path, data_only=True)

    # Normalize sheet names
    sheets = {s.lower().replace(" ", ""): s for s in wb.sheetnames}
    get_sheet = lambda name: sheets.get(name.lower().replace(" ", ""))

    runs_sheet = wb[get_sheet("TableRuns")]
    tables_sheet = wb[get_sheet("Tables")]
    pops_sheet = wb[get_sheet("Populations")]
    params_sheet = wb[get_sheet("Parameters")]

    # Extract Tablepgm, PopID, ParamID from TableRuns
    tablepgm = popid = paramid = ""
    for row in runs_sheet.iter_rows(min_row=2, values_only=True):
        tid = str(row[1]).strip()
        if tid == table_id:
            tablepgm = str(row[4]).strip() if row[4] else ""
            popid = str(row[6]).strip() if row[6] else ""
            paramid = str(row[9]).strip() if row[9] else ""
            break

    print(f"✅ ParamID from Table Runs: {paramid}")

    # Get ParameterLabel
    param_label = ""
    for row in params_sheet.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip() == paramid:
            param_label = str(row[3]).strip() if row[3] else ""
            break

    print(f"✅ ParameterLabel from Parameters sheet: {param_label}")

    # Extract Titles from Tables sheet
    title1_parts = []
    for row in tables_sheet.iter_rows(min_row=2, values_only=True):
        pgm = str(row[3]).strip() if row[3] else ""
        if pgm == tablepgm:
            title1_parts = [row[9], row[10], row[11]]  # Columns J, K, L
            break

    title_line1 = f"Table {table_id}"
    title_line2 = " - ".join([substitute(str(t), param_label) for t in title1_parts if t])
    title_line3 = ""

    for row in pops_sheet.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip() == popid:
            title_line3 = str(row[2]).strip() if row[2] else ""
            break

    print("✅ Extracted Titles:")
    print("🔹 Title Line 1:", title_line1)
    print("🔹 Title Line 2:", title_line2)
    print("🔹 Title Line 3:", title_line3)

# Example usage
excel_path = "P:/Biostatistics/PDF Tool/Shell Generator/Training March 2022 TLFs.xlsx"
tid = "14.1.5"

get_titles_with_substitution(excel_path, tid)
