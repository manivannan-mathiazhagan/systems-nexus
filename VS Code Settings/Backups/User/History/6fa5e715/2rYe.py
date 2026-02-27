import os
from openpyxl import load_workbook

excel_path = "P:/Biostatistics/PDF Tool/Shell Generator/Training March 2022 TLFs.xlsx"
tid = "14.1.5"

def clean(val):
    return str(val).strip() if val is not None else ""

def get_sheet_by_normalized_name(wb, target_name):
    sheets = {s.lower().replace(" ", ""): s for s in wb.sheetnames}
    return wb[sheets.get(target_name.lower().replace(" ", ""))]

wb = load_workbook(excel_path, data_only=True)

# Sheets
runs = get_sheet_by_normalized_name(wb, "Table Runs")
tables = get_sheet_by_normalized_name(wb, "Tables")
pops = get_sheet_by_normalized_name(wb, "Populations")
params = get_sheet_by_normalized_name(wb, "Parameters")

# Get TNumManual, pgm, popid, paramid for selected table ID
pgm = popid = paramid = tnum = None
for row in runs.iter_rows(min_row=2, values_only=True):
    if clean(row[1]) == tid:
        tnum = clean(row[1])  # TNumManual
        pgm = clean(row[4])   # Tablepgm
        popid = clean(row[6]) # PopID
        paramid = clean(row[9])  # ParamID
        break

if not pgm:
    print(f"❌ Table ID {tid} not found in Table Runs.")
    exit()

# Get Title 2 (from Tables)
title2 = ""
for row in tables.iter_rows(min_row=2, values_only=True):
    if clean(row[3]) == pgm:
        parts = [clean(row[9]), clean(row[10]), clean(row[11])]  # Title1, Title2, Title3
        title2 = " - ".join([p for p in parts if p])
        break

# Get Title 3 (from Populations)
title3 = ""
for row in pops.iter_rows(min_row=2, values_only=True):
    if clean(row[0]) == popid:
        title3 = clean(row[2])  # Population Name
        break

# Get Parameter Label from Parameters
param_label = ""
for row in params.iter_rows(min_row=2, values_only=True):
    if clean(row[0]) == paramid:
        param_label = clean(row[3])  # Column 4 is ParameterLabel
        break

# Substitution
def substitute(text, label):
    return text.replace("&ParameterLabel", label) if text else ""

title1 = f"Table {tnum}"
title2 = substitute(title2, param_label)
title3 = substitute(title3, param_label)

print("✅ Extracted Titles:")
print(f"🔹 Title Line 1: {title1}")
print(f"🔹 Title Line 2: {title2}")
print(f"🔹 Title Line 3: {title3}")
