from openpyxl import load_workbook

excel_path = "P:/Biostatistics/PDF Tool/Shell Generator/Training March 2022 TLFs.xlsx"  # Replace with your actual path
tid = "14.1.5"

wb = load_workbook(excel_path, data_only=True)

# Get required sheets
runs = wb["Table Runs"]
params = wb["Parameters"]
tables = wb["Tables"]

# Extract ParamID for the given Table ID
paramid = None
for row in runs.iter_rows(min_row=2, values_only=True):
    if str(row[1]).strip() == tid:
        paramid = str(row[9]).strip()  # Assuming column 10 = ParamIDs
        break

print(f"✅ ParamID from Table Runs: {paramid}")

# Extract ParameterLabel for that ParamID from column 4 of Parameters sheet
paramlabel = None
for row in params.iter_rows(min_row=2, values_only=True):
    if str(row[0]).strip() == paramid:
        paramlabel = str(row[3]).strip()  # Column 4 = ParameterLabel
        break

print(f"✅ ParameterLabel from Parameters sheet: {paramlabel}")

# Fetch Title Line 2 from Tables sheet (should contain &ParameterLabel)
title2 = ""
for row in tables.iter_rows(min_row=2, values_only=True):
    if str(row[3]).strip() == "t14_1_5":  # Assuming Tablepgm = t14_1_5 (column 4)
        titles = [t for t in (row[9], row[10], row[11]) if t]
        title2 = " - ".join(str(t) for t in titles)
        break

print(f"🔤 Raw Title Line 2: {title2}")
print(f"✅ After substitution: {title2.replace('&ParameterLabel', paramlabel)}")
