import sys, os
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QTextEdit, QFileDialog, QLineEdit, QMessageBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QFrame
)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
import win32com.client as win32


class TableShellTool(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("📄 Table Shell Exporter")
        self.resize(1200, 800)
        self.excel_data = {}
        self.param_labels = {}
        self.title_2_map = {}
        self.title_3_map = {}
        self.trt_group_map = {}
        self.current_table_id = ""
        self.header_footer_data = {"Header": {}, "Footer": {}}
        self.header_footer_visible = True
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()
        layout.addWidget(QLabel("<h2>📘 Clinical Table Shell Builder</h2>"))

        file_row = QHBoxLayout()
        self.excel_path = QLineEdit()
        self.excel_path.setReadOnly(True)
        browse_btn = QPushButton("Browse Excel Spec")
        browse_btn.clicked.connect(self.load_excel)
        file_row.addWidget(self.excel_path)
        file_row.addWidget(browse_btn)
        layout.addLayout(file_row)

        layout.addWidget(QLabel("Select Table ID:"))
        self.table_dropdown = QComboBox()
        self.table_dropdown.currentTextChanged.connect(self.load_table_content)
        layout.addWidget(self.table_dropdown)

        self.toggle_header_footer_btn = QPushButton("🔽 Show/Hide Header/Footer")
        self.toggle_header_footer_btn.clicked.connect(self.toggle_header_footer)
        layout.addWidget(self.toggle_header_footer_btn)

        self.header_footer_container = QFrame()
        hf_layout = QVBoxLayout()
        self.header_preview = QTextEdit()
        self.footer_preview = QTextEdit()
        self.header_preview.setReadOnly(True)
        self.footer_preview.setReadOnly(True)
        hf_layout.addWidget(QLabel("Header Preview:"))
        hf_layout.addWidget(self.header_preview)
        hf_layout.addWidget(QLabel("Footer Preview:"))
        hf_layout.addWidget(self.footer_preview)
        self.header_footer_container.setLayout(hf_layout)
        layout.addWidget(self.header_footer_container)

        for i, label in enumerate(["Title Line 1", "Title Line 2", "Title Line 3"]):
            layout.addWidget(QLabel(label))
            te = QTextEdit()
            te.setFixedHeight(30)
            setattr(self, f"title{i+1}_edit", te)
            layout.addWidget(te)

        self.grid = QTableWidget(0, 3)
        self.grid.setHorizontalHeaderLabels(["Shell Row Text", "Format", "Dummy Value"])
        self.grid.setEditTriggers(QAbstractItemView.AllEditTriggers)
        self.grid.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.grid)

        add_btn = QPushButton("➕ Add Row")
        add_btn.clicked.connect(self.add_row)
        layout.addWidget(add_btn)

        export_row = QHBoxLayout()
        for label, method in [("📄 Export RTF", self.export_rtf),
                              ("📄 Export PDF", self.export_pdf),
                              ("💻 Export SAS Macro", self.export_sas)]:
            btn = QPushButton(label)
            btn.clicked.connect(method)
            export_row.addWidget(btn)
        layout.addLayout(export_row)

        self.status = QLabel("Ready.")
        layout.addWidget(self.status)
        self.setLayout(layout)

    def toggle_header_footer(self):
        self.header_footer_visible = not self.header_footer_visible
        self.header_footer_container.setVisible(self.header_footer_visible)
            
    def load_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel Spec", "", "Excel Files (*.xlsx *.xls)")
        if not path:
            return
        self.excel_path.setText(path)
        wb = load_workbook(path, data_only=True)

        sheets = {s.lower().replace(" ", ""): s for s in wb.sheetnames}
        get_sheet = lambda name: sheets.get(name.lower().replace(" ", ""))

        runs = wb[get_sheet("TableRuns")]
        tables = wb[get_sheet("Tables")]
        pops = wb[get_sheet("Populations")]
        header_sheet = get_sheet("HeadersFooters") or get_sheet("HeadersandFooters")
        headers = wb[header_sheet] if header_sheet else None
        params = wb[get_sheet("Parameters")]

        self.excel_data.clear()
        self.table_dropdown.clear()
        self.header_footer_data = {"Header": {}, "Footer": {}}
        self.param_labels.clear()
        self.title_2_map.clear()
        self.title_3_map.clear()

        for row in runs.iter_rows(min_row=2, values_only=True):
            tid, pgm, popid, paramids = row[1], row[4], row[6], row[9]
            if tid and pgm:
                tid, pgm, popid = str(tid).strip(), str(pgm).strip(), str(popid or "").strip()
                self.excel_data[tid] = {"pgm": pgm, "popid": popid, "paramids": str(paramids or "")}
                self.table_dropdown.addItem(tid)

        for row in tables.iter_rows(min_row=2, values_only=True):
            pgm = str(row[3]).strip() if row[3] else None
            titles = [str(row[9] or ""), str(row[10] or ""), str(row[11] or "")]
            if pgm:
                self.title_2_map[pgm] = titles  # store as list of three title lines

        for row in pops.iter_rows(min_row=2, values_only=True):
            pid, name = str(row[0]).strip(), str(row[2]).strip()
            self.title_3_map[pid] = name

        for row in params.iter_rows(min_row=2, values_only=True):
            pid = str(row[0]).strip()
            label = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            self.param_labels[pid] = label

        if headers:
            for r in headers.iter_rows(min_row=2, values_only=True):
                try:
                    section = r[0]
                    line = int(r[1])
                    if section in ["Header", "Footer"]:
                        self.header_footer_data[section].setdefault(line, ["", "", ""])
                        for i in range(3):
                            if r[i+2]:
                                self.header_footer_data[section][line][i] = str(r[i+2]).strip()
                except:
                    continue
        def load_table_content(self, tid):
        self.current_table_id = tid
        self.grid.setRowCount(0)

        data = self.excel_data.get(tid, {})
        pgm = data.get("pgm", "")
        popid = data.get("popid", "")
        paramids = data.get("paramids", "")
        param_label = self.param_labels.get(paramids.strip(), "") if paramids else ""

        def sub_title(text):
            if not text:
                return ""
            return text.replace("&ParameterLabel", param_label)

        raw_titles = self.title_2_map.get(pgm, ["", "", ""])
        raw_title1 = f"Table {tid}"
        raw_title2 = raw_titles[0] if len(raw_titles) > 0 else ""
        raw_title3 = self.title_3_map.get(popid, "")

        self.title1_edit.setText(sub_title(raw_title1))
        self.title2_edit.setText(sub_title(raw_title2))
        self.title3_edit.setText(sub_title(raw_title3))

        self.render_header_footer_preview()
        self.header_footer_container.setVisible(self.header_footer_visible)

        self.grid.setRowCount(3)
        for r in range(3):
            self.grid.setItem(r, 0, QTableWidgetItem(f"Row {r+1}"))
            fmt = QComboBox()
            fmt.addItems(["n", "n(%)", "Mean(SD)", "Min", "Max", "NPCT"])
            self.grid.setCellWidget(r, 1, fmt)
            self.grid.setItem(r, 2, QTableWidgetItem("xx"))
    def export_pdf(self):
        rtf_path, _ = QFileDialog.getOpenFileName(self, "Select RTF", "", "RTF Files (*.rtf)")
        if not rtf_path:
            return
        pdf_path, _ = QFileDialog.getSaveFileName(self, "Save PDF", rtf_path.replace(".rtf", ".pdf"), "PDF Files (*.pdf)")
        if not pdf_path:
            return
        try:
            word = win32.gencache.EnsureDispatch("Word.Application")
            doc = word.Documents.Open(rtf_path)
            doc.SaveAs(pdf_path, FileFormat=17)
            doc.Close()
            word.Quit()
            self.status.setText(f"✅ PDF exported: {pdf_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"PDF export failed:\n{e}")

    def export_sas(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save SAS", f"{self.current_table_id}_macro.sas", "SAS Files (*.sas)")
        if not path:
            return
        try:
            code = f"%macro table_{self.current_table_id.replace('.', '_')}();\n"
            for i in range(1, 4):
                title = getattr(self, f"title{i}_edit").toPlainText().strip()
                if title:
                    code += f"title{i} '{title}';\n"
            code += "proc report data=mydata nowd;\ncolumns "
            for r in range(self.grid.rowCount()):
                col = self.grid.item(r, 0)
                if col:
                    text = col.text().strip().replace(" ", "_")
                    code += f"{text} "
            code += ";\nrun;\n%mend;\n"
            with open(path, "w") as f:
                f.write(code)
            self.status.setText(f"✅ SAS macro saved: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"SAS Export failed:\n{e}")
            
if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = TableShellTool()
    w.show()
    sys.exit(app.exec_())

