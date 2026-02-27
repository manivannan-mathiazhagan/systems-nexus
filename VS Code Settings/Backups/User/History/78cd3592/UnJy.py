import sys, os
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QTextEdit, QFileDialog, QLineEdit, QMessageBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QFrame
)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
import win32com.client as win32


class TableShellTool(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("📄 Table Shell Exporter v4.1")
        self.resize(1200, 800)
        self.excel_data = {}
        self.title_2_map = {}
        self.title_3_map = {}
        self.trt_group_map = {}
        self.spanning_headers = {1: [], 2: [], 3: []}
        self.current_table_id = ""
        self.header_footer_data = {"Header": {}, "Footer": {}}
        self.header_footer_visible = True
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()
        layout.addWidget(QLabel("<h2>📘 Clinical Table Shell Builder</h2>"))

        file_row = QHBoxLayout()
        self.excel_path = QLineEdit()
        self.excel_path.setReadOnly(True)
        browse_btn = QPushButton("Browse Excel Spec")
        browse_btn.clicked.connect(self.load_excel)
        file_row.addWidget(self.excel_path)
        file_row.addWidget(browse_btn)
        layout.addLayout(file_row)

        layout.addWidget(QLabel("Select Table ID:"))
        self.table_dropdown = QComboBox()
        self.table_dropdown.currentTextChanged.connect(self.load_table_content)
        layout.addWidget(self.table_dropdown)

        self.toggle_header_footer_btn = QPushButton("🔽 Show/Hide Header/Footer")
        self.toggle_header_footer_btn.clicked.connect(self.toggle_header_footer)
        layout.addWidget(self.toggle_header_footer_btn)

        self.header_footer_container = QFrame()
        hf_layout = QVBoxLayout()
        self.header_preview = QTextEdit()
        self.footer_preview = QTextEdit()
        self.header_preview.setReadOnly(True)
        self.footer_preview.setReadOnly(True)
        hf_layout.addWidget(QLabel("Header Preview:"))
        hf_layout.addWidget(self.header_preview)
        hf_layout.addWidget(QLabel("Footer Preview:"))
        hf_layout.addWidget(self.footer_preview)
        self.header_footer_container.setLayout(hf_layout)
        layout.addWidget(self.header_footer_container)

        for i, label in enumerate(["Title Line 1", "Title Line 2", "Title Line 3"]):
            layout.addWidget(QLabel(label))
            te = QTextEdit()
            te.setFixedHeight(30)
            setattr(self, f"title{i+1}_edit", te)
            layout.addWidget(te)

        self.grid = QTableWidget(0, 3)
        self.grid.setHorizontalHeaderLabels(["Shell Row Text", "Format", ""])
        self.grid.setEditTriggers(QAbstractItemView.AllEditTriggers)
        self.grid.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.grid)

        add_btn = QPushButton("➕ Add Row")
        add_btn.clicked.connect(self.add_row)
        layout.addWidget(add_btn)

        export_row = QHBoxLayout()
        for label, method in [("📄 Export RTF", self.export_rtf),
                              ("📄 Export PDF", self.export_pdf),
                              ("💻 Export SAS Macro", self.export_sas)]:
            btn = QPushButton(label)
            btn.clicked.connect(method)
            export_row.addWidget(btn)
        layout.addLayout(export_row)

        self.status = QLabel("Ready.")
        layout.addWidget(self.status)
        self.setLayout(layout)

    def toggle_header_footer(self):
        self.header_footer_visible = not self.header_footer_visible
        self.header_footer_container.setVisible(self.header_footer_visible)

    def load_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel Spec", "", "Excel Files (*.xlsx *.xls)")
        if not path:
            return
        self.excel_path.setText(path)
        wb = load_workbook(path, data_only=True)

        sheets = {s.lower().replace(" ", ""): s for s in wb.sheetnames}
        get_sheet = lambda name: sheets.get(name.lower().replace(" ", ""))

        runs_sheet = get_sheet("TableRuns")
        tables_sheet = get_sheet("Tables")
        pops_sheet = get_sheet("Populations")
        trts_sheet = get_sheet("TreatmentGroups")
        spans_sheet = get_sheet("SpanningHeaders")
        header_sheet = get_sheet("HeadersFooters") or get_sheet("HeadersandFooters")

        self.excel_data.clear()
        self.trt_group_map.clear()
        self.header_footer_data = {"Header": {}, "Footer": {}}
        self.table_dropdown.clear()

        ws = wb[runs_sheet]
        table_info = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            tid, pgm, popid = row[1], row[4], row[6]
            if tid and pgm and "x" not in str(tid).lower():
                table_info.append((str(tid).strip(), str(pgm).strip(), str(popid).strip()))
        for tid, pgm, popid in table_info:
            self.excel_data[tid] = {"pgm": pgm, "popid": popid}
            self.table_dropdown.addItem(tid)

        ws = wb[tables_sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            pgm = str(row[3]).strip() if row[3] else None
            titles = [t for t in (row[9], row[10], row[11]) if t]
            if pgm:
                self.title_2_map[pgm] = " - ".join(str(t) for t in titles if t)

        ws = wb[pops_sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            pid, name = str(row[0]).strip(), str(row[2]).strip()
            self.title_3_map[pid] = name

        ws = wb[trts_sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            gid, label = str(row[0]).strip(), str(row[1]).strip()
            self.trt_group_map.setdefault(gid, []).append(label)

        if spans_sheet:
            ws = wb[spans_sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    level = int(row[0])
                    self.spanning_headers.setdefault(level, []).append(
                        ";".join(str(cell) for cell in row[1:] if cell)
                    )
                except:
                    continue

        if header_sheet:
            ws = wb[header_sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                section = row[0]
                line = row[1]
                if section in ["Header", "Footer"] and isinstance(line, int):
                    self.header_footer_data[section][line] = [
                        str(row[2] or ""), str(row[3] or ""), str(row[4] or "")
                    ]

    def load_table_content(self, table_id):
        self.current_table_id = table_id
        self.grid.setRowCount(0)

        pgm = self.excel_data.get(table_id, {}).get("pgm", "").strip()
        popid = self.excel_data.get(table_id, {}).get("popid", "").strip()

        self.title1_edit.setText(f"Table {table_id}")
        self.title2_edit.setText(self.title_2_map.get(pgm, "") or "")
        self.title3_edit.setText(self.title_3_map.get(popid, "") or "")

        self.render_header_footer_preview()
        self.header_footer_container.setVisible(self.header_footer_visible)

        self.grid.setRowCount(3)
        for r in range(3):
            self.grid.setItem(r, 0, QTableWidgetItem(f"Row {r+1}"))
            fmt = QComboBox()
            fmt.addItems(["n", "n(%)", "Mean(SD)", "Min", "Max", "NPCT"])
            self.grid.setCellWidget(r, 1, fmt)
            self.grid.setItem(r, 2, QTableWidgetItem("xx"))

    def render_header_footer_preview(self):
        def build_text(data):
            lines = []
            for i in sorted(data.keys()):
                left, center, right = data[i]
                lines.append(f"{left:<40}{center:^40}{right:>40}".rstrip())
            return "\n".join(lines)

        self.header_preview.setText(build_text(self.header_footer_data["Header"]))
        self.footer_preview.setText(build_text(self.header_footer_data["Footer"]))

    def add_row(self):
        row = self.grid.rowCount()
        self.grid.insertRow(row)
        self.grid.setItem(row, 0, QTableWidgetItem(""))
        combo = QComboBox()
        combo.addItems(["n", "n(%)", "Mean(SD)", "Min", "Max", "NPCT"])
        self.grid.setCellWidget(row, 1, combo)
        self.grid.setItem(row, 2, QTableWidgetItem("xx"))

    def export_rtf(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save RTF File", f"{self.current_table_id}.rtf", "RTF Files (*.rtf)")
        if not path:
            return
        try:
            rtf = [
                r"{\rtf1\ansi\deff0",
                r"{\fonttbl{\f0 Courier New;}}",
                r"\fs16",  # 8pt font
                r"\paperw16840\paperh11900\landscape",
                r"\margl1440\margr1440\margt1440\margb1440"
            ]

            def build_hf(lines):
                result = []
                for i in sorted(lines.keys()):
                    l, c, r = lines[i]
                    result.append(r"""
\trowd\trgaph108\trleft-108
\cellx4000\cellx8000\cellx12000
\pard\plain\ql\f0\fs16 """ + l + r"""\cell
\pard\plain\qc\f0\fs16 """ + c + r"""\cell
\pard\plain\qr\f0\fs16 """ + r + r"""\cell\row""")
                return "".join(result)

            rtf.append(r"{\header " + build_hf(self.header_footer_data["Header"]) + "}")
            rtf.append(r"{\footer " + build_hf(self.header_footer_data["Footer"]) + "}")

            for i in range(1, 4):
                title = getattr(self, f"title{i}_edit").toPlainText().strip()
                if title:
                    rtf.append(rf"\pard\qc\f0\fs16 {title}\par")

            rtf.append(r"\trowd\trgaph108\trleft-108")
            rtf.append(r"\cellx3000\cellx6000\cellx9000")

            for r in range(self.grid.rowCount()):
                text = self.grid.item(r, 0).text() if self.grid.item(r, 0) else ""
                fmt = self.grid.cellWidget(r, 1).currentText()
                val = self.grid.item(r, 2).text() if self.grid.item(r, 2) else ""
                rtf.append(rf"\intbl\f0\fs16 {text}\cell {fmt}\cell {val}\cell\row")

            rtf.append("}")

            with open(path, "w") as f:
                f.write("\n".join(rtf))

            self.status.setText(f"✅ RTF exported: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"RTF export failed:\n{str(e)}")

    def export_pdf(self):
        rtf_path, _ = QFileDialog.getOpenFileName(self, "Select RTF File", "", "RTF Files (*.rtf)")
        if not rtf_path:
            return
        pdf_path, _ = QFileDialog.getSaveFileName(self, "Save PDF File", rtf_path.replace(".rtf", ".pdf"), "PDF Files (*.pdf)")
        if not pdf_path:
            return
        try:
            word = win32.gencache.EnsureDispatch("Word.Application")
            doc = word.Documents.Open(rtf_path)
            doc.SaveAs(pdf_path, FileFormat=17)
            doc.Close()
            word.Quit()
            self.status.setText(f"✅ PDF exported: {pdf_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"PDF export failed:\n{str(e)}")

    def export_sas(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save SAS Macro", f"{self.current_table_id}_macro.sas", "SAS Files (*.sas)")
        if not path:
            return
        try:
            code = f"%macro table_{self.current_table_id.replace('.', '_')}();\n"
            for i in range(1, 4):
                title = getattr(self, f"title{i}_edit").toPlainText().strip()
                if title:
                    code += f"title{i} '{title}';\n"
            code += "proc report data=mydata nowd;\ncolumns "

            for r in range(self.grid.rowCount()):
                col = self.grid.item(r, 0)
                if col:
                    text = col.text().strip().replace(" ", "_")
                    code += f"{text} "
            code += ";\nrun;\n%mend;\n"

            with open(path, "w") as f:
                f.write(code)

            self.status.setText(f"✅ SAS macro saved: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"SAS export failed:\n{str(e)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = TableShellTool()
    window.show()
    sys.exit(app.exec_())
