import sys
import os
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QFileDialog, QMessageBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QFrame, QTextEdit, QScrollArea
)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
import win32com.client as win32


class TableShellTool(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("📄 Table Shell Generator")
        self.resize(1200, 800)
        self.excel_data = {}
        self.current_table_id = ""
        self.header_footer_data = {"Header": {}, "Footer": {}}
        self.param_labels = {}
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()

        file_layout = QHBoxLayout()
        self.file_label = QLabel("No Excel file selected")
        file_button = QPushButton("Load Excel Spec")
        file_button.clicked.connect(self.load_excel)
        file_layout.addWidget(self.file_label)
        file_layout.addWidget(file_button)

        table_layout = QHBoxLayout()
        self.table_dropdown = QComboBox()
        self.table_dropdown.currentTextChanged.connect(self.load_table_data)
        table_layout.addWidget(QLabel("Select Table ID:"))
        table_layout.addWidget(self.table_dropdown)

        self.table_preview = QTableWidget()
        self.table_preview.setEditTriggers(QTableWidget.DoubleClicked)
        self.table_preview.setSelectionBehavior(QTableWidget.SelectItems)
        self.table_preview.horizontalHeader().setStretchLastSection(True)
        self.table_preview.verticalHeader().setVisible(False)
        self.table_preview.setAlternatingRowColors(True)
        self.table_preview.setWordWrap(False)

        export_layout = QHBoxLayout()
        export_btn = QPushButton("Export RTF")
        export_btn.clicked.connect(self.export_rtf)
        export_pdf_btn = QPushButton("Export PDF")
        export_pdf_btn.clicked.connect(self.export_pdf)
        export_layout.addWidget(export_btn)
        export_layout.addWidget(export_pdf_btn)

        layout.addLayout(file_layout)
        layout.addLayout(table_layout)
        layout.addWidget(self.table_preview)
        layout.addLayout(export_layout)

        self.setLayout(layout)

    def load_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        self.file_label.setText(os.path.basename(file_path))
        wb = load_workbook(file_path, data_only=True)

        try:
            self.excel_data = {
                "Tables": {row[3].value: row for row in wb["Tables"].iter_rows(min_row=2) if row[3].value},
                "Table Runs": [row for row in wb["Table Runs"].iter_rows(min_row=2)],
                "Populations": {str(row[0].value): row[2].value for row in wb["Populations"].iter_rows(min_row=2) if row[0].value},
                "Parameters": {str(row[0].value): row[3].value for row in wb["Parameters"].iter_rows(min_row=2) if row[0].value},
                "Headersfooters": self.parse_header_footer(wb)
            }
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error loading Excel data:\n{e}")
            return

        table_ids = list({row[1].value for row in self.excel_data["Table Runs"] if row[1].value})
        self.table_dropdown.clear()
        self.table_dropdown.addItems(sorted(table_ids))

    def parse_header_footer(self, wb):
        for name in ["Headersfooters", "Headers and Footers"]:
            if name in wb:
                sheet = wb[name]
                output = {"Header": {}, "Footer": {}}
                for row in sheet.iter_rows(min_row=2):
                    kind = str(row[0].value).strip().capitalize()
                    try:
                        line = int(row[1].value)
                    except:
                        continue
                    if kind not in output:
                        continue
                    left = row[2].value or ""
                    center = row[3].value or ""
                    right = row[4].value or ""
                    output[kind].setdefault(line, []).append((left, center, right))
                return output
        return {"Header": {}, "Footer": {}}

    def load_table_data(self, table_id):
        self.current_table_id = table_id
        self.table_preview.clear()
        self.table_preview.setRowCount(0)
        self.table_preview.setColumnCount(1)
        self.table_preview.setHorizontalHeaderLabels(["Content"])

        rows = []

        runs = [r for r in self.excel_data["Table Runs"] if str(r[1].value).strip() == str(table_id).strip()]
        if not runs:
            return

        run = runs[0]
        tnum = run[1].value
        pgm = run[3].value
        popid = str(run[4].value).strip() if run[4].value else ""
        paramid = str(run[5].value).strip() if run[5].value else ""

        table_row = self.excel_data["Tables"].get(pgm, [])
        title_parts = [table_row[9].value, table_row[10].value, table_row[11].value] if table_row else ["", "", ""]
        title_lines = [f"Table {tnum}"] + [t for t in title_parts if t]

        if paramid and "&ParameterLabel" in " ".join(title_lines):
            label = self.excel_data["Parameters"].get(paramid, "")
            title_lines = [t.replace("&ParameterLabel", label) for t in title_lines]

        if popid:
            popname = self.excel_data["Populations"].get(popid, "")
            if popname:
                title_lines.append(popname)

        for line_num in sorted(self.excel_data["Headersfooters"]["Header"]):
            for left, center, right in self.excel_data["Headersfooters"]["Header"][line_num]:
                line = f"{left:<40}{center:^40}{right:>40}".rstrip()
                rows.append(line)

        for title in title_lines:
            rows.append(f"{title:^120}")

        rows.append("-" * 120)
        rows.append("Sample table content goes here".center(120))
        rows.append("-" * 120)

        for line_num in sorted(self.excel_data["Headersfooters"]["Footer"]):
            for left, center, right in self.excel_data["Headersfooters"]["Footer"][line_num]:
                line = f"{left:<40}{center:^40}{right:>40}".rstrip()
                rows.append(line)

        self.table_preview.setRowCount(len(rows))
        for i, row_text in enumerate(rows):
            item = QTableWidgetItem(row_text)
            self.table_preview.setItem(i, 0, item)

    def export_rtf(self):
        if not self.current_table_id:
            QMessageBox.warning(self, "Warning", "No table selected.")
            return

        save_path, _ = QFileDialog.getSaveFileName(self, "Save RTF", f"{self.current_table_id}.rtf", "RTF Files (*.rtf)")
        if not save_path:
            return

        try:
            with open(save_path, "w") as f:
                f.write("{\\rtf1\\ansi\\landscape\\deff0\n")
                f.write("{\\fonttbl{\\f0 Courier New;}}\n")
                f.write("\\fs16\n")  # 8 pt

                for i in range(self.table_preview.rowCount()):
                    text = self.table_preview.item(i, 0).text()
                    f.write(f"{text}\\line\n")

                f.write("}")
            QMessageBox.information(self, "Success", f"RTF saved to:\n{save_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to write RTF:\n{e}")

    def export_pdf(self):
        if not self.current_table_id:
            QMessageBox.warning(self, "Warning", "No table selected.")
            return

        rtf_path, _ = QFileDialog.getSaveFileName(self, "Save PDF", f"{self.current_table_id}.pdf", "PDF Files (*.pdf)")
        if not rtf_path:
            return

        rtf_temp = rtf_path.replace(".pdf", ".rtf")
        self.export_rtf_internal(rtf_temp)

        try:
            word = win32.gencache.EnsureDispatch('Word.Application')
            doc = word.Documents.Open(rtf_temp)
            doc.SaveAs(rtf_path, FileFormat=17)
            doc.Close()
            word.Quit()
            os.remove(rtf_temp)
            QMessageBox.information(self, "Success", f"PDF saved to:\n{rtf_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"PDF conversion failed:\n{e}")

    def export_rtf_internal(self, save_path):
        try:
            with open(save_path, "w") as f:
                f.write("{\\rtf1\\ansi\\landscape\\deff0\n")
                f.write("{\\fonttbl{\\f0 Courier New;}}\n")
                f.write("\\fs16\n")  # 8 pt

                for i in range(self.table_preview.rowCount()):
                    text = self.table_preview.item(i, 0).text()
                    f.write(f"{text}\\line\n")

                f.write("}")
        except:
            pass


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = TableShellTool()
    window.show()
    sys.exit(app.exec_())
