import sys, os, re
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QTextEdit, QFileDialog, QLineEdit, QMessageBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QFrame
)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
import win32com.client as win32

class TableShellTool(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("📄 Table Shell Exporter")
        self.resize(1200, 900)
        self.excel_data = {}
        self.param_labels = {}
        self.param_foot = {}
        self.title_2_map = {}
        self.table_foot = {}
        self.title_3_map = {}
        self.current_table_id = ""
        self.header_footer_data = {"Header": {}, "Footer": {}}
        self.header_footer_visible = True
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()
        layout.addWidget(QLabel("<h2>📘 Clinical Table Shell Builder</h2>"))

        file_row = QHBoxLayout()
        self.excel_path = QLineEdit(); self.excel_path.setReadOnly(True)
        browse_btn = QPushButton("Browse Excel Spec"); browse_btn.clicked.connect(self.load_excel)
        file_row.addWidget(self.excel_path); file_row.addWidget(browse_btn)
        layout.addLayout(file_row)

        layout.addWidget(QLabel("Select Table ID:"))
        self.table_dropdown = QComboBox()
        self.table_dropdown.currentTextChanged.connect(self.load_table_content)
        layout.addWidget(self.table_dropdown)

        self.toggle_header_footer_btn = QPushButton("🔽 Show/Hide Header/Footer")
        self.toggle_header_footer_btn.clicked.connect(self.toggle_header_footer)
        layout.addWidget(self.toggle_header_footer_btn)

        self.header_footer_container = QFrame()
        hf_layout = QVBoxLayout()
        self.header_preview = QTextEdit(); self.footer_preview = QTextEdit()
        self.header_preview.setReadOnly(True); self.footer_preview.setReadOnly(True)
        hf_layout.addWidget(QLabel("Header Preview:")); hf_layout.addWidget(self.header_preview)
        hf_layout.addWidget(QLabel("Footer Preview:")); hf_layout.addWidget(self.footer_preview)
        self.header_footer_container.setLayout(hf_layout)
        layout.addWidget(self.header_footer_container)

        for i, label in enumerate(["Title Line 1", "Title Line 2", "Title Line 3"]):
            layout.addWidget(QLabel(label))
            te = QTextEdit(); te.setFixedHeight(30)
            setattr(self, f"title{i+1}_edit", te)
            layout.addWidget(te)

        layout.addWidget(QLabel("Footnotes:"))
        self.footnote_box = QTextEdit()
        self.footnote_box.setReadOnly(True)
        self.footnote_box.setFixedHeight(60)
        layout.addWidget(self.footnote_box)

        self.grid = QTableWidget(0, 3)
        self.grid.setHorizontalHeaderLabels(["Shell Row Text", "Format", "Dummy Value"])
        self.grid.setEditTriggers(QAbstractItemView.AllEditTriggers)
        self.grid.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.grid)

        add_btn = QPushButton("➕ Add Row"); add_btn.clicked.connect(self.add_row)
        layout.addWidget(add_btn)

        export_row = QHBoxLayout()
        for label, method in [("📄 Export RTF", self.export_rtf),
                              ("📄 Export PDF", self.export_pdf),
                              ("💻 Export SAS Macro", self.export_sas)]:
            btn = QPushButton(label); btn.clicked.connect(method)
            export_row.addWidget(btn)
        layout.addLayout(export_row)

        self.status = QLabel("Ready."); layout.addWidget(self.status)
        self.setLayout(layout)

    def toggle_header_footer(self):
        self.header_footer_visible = not self.header_footer_visible
        self.header_footer_container.setVisible(self.header_footer_visible)

    def load_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel Spec", "", "Excel Files (*.xlsx *.xls)")
        if not path: return
        self.excel_path.setText(path)
        wb = load_workbook(path, data_only=True)

        sheets = {s.lower().replace(" ", ""): s for s in wb.sheetnames}
        get_sheet = lambda name: sheets.get(name.lower().replace(" ", ""))

        runs = wb[get_sheet("TableRuns")]
        tables = wb[get_sheet("Tables")]
        pops = wb[get_sheet("Populations")]
        headers = wb[get_sheet("HeadersFooters") or get_sheet("HeadersandFooters")]
        params = wb[get_sheet("Parameters")]

        self.excel_data.clear(); self.table_dropdown.clear()
        self.param_labels.clear(); self.param_foot.clear()
        self.title_2_map.clear(); self.table_foot.clear()
        self.title_3_map.clear(); self.header_footer_data = {"Header":{}, "Footer":{}}
        self.footnote_box.clear()

        for row in runs.iter_rows(min_row=2, values_only=True):
            tid, pgm, popid, paramids = row[1], row[4], row[6], row[9]
            if tid and pgm:
                tid, pgm, popid = str(tid).strip(), str(pgm).strip(), str(popid or "").strip()
                self.excel_data[tid] = {"pgm": pgm, "popid": popid, "paramids": str(paramids or "")}
                self.table_dropdown.addItem(tid)

        for r in tables.iter_rows(min_row=2, values_only=True):
            pgm = str(r[3]).strip() if r[3] else ""
            tbl_foot = str(r[8] or "").strip()
            tbl_titles = [str(r[9] or ""), str(r[10] or ""), str(r[11] or "")]
            if pgm:
                self.title_2_map[pgm] = tbl_titles
                self.table_foot[pgm] = tbl_foot

        for r in params.iter_rows(min_row=2, values_only=True):
            pid = str(r[0]).strip()
            self.param_labels[pid] = str(r[2] or "")
            self.param_foot[pid] = str(r[3] or "")

        for r in pops.iter_rows(min_row=2, values_only=True):
            pid, name = str(r[0]).strip(), str(r[2] or "").strip()
            self.title_3_map[pid] = name

        for r in headers.iter_rows(min_row=2, values_only=True):
            section, line = r[0], r[1]
            if section in ("Header", "Footer") and isinstance(line, (int, float)):
                self.header_footer_data[section].setdefault(int(line), ["","",""])
                for i in range(3):
                    if r[2+i]:
                        self.header_footer_data[section][int(line)][i] = str(r[2+i]).strip()

    def load_table_content(self, tid):
        self.current_table_id = tid
        self.grid.setRowCount(0)

        row = self.excel_data[tid]
        pgm, popid, pid = row["pgm"], row["popid"], row["paramids"]
        param_label = self.param_labels.get(pid, "")
        tbl_foot = self.table_foot.get(pgm, "")
        param_foot = self.param_foot.get(pid, "")

        def subs(txt):
            return re.sub(r"&parameterlabel", param_label, txt or "", flags=re.IGNORECASE)

        raw1 = f"Table {tid}"
        raw2 = self.title_2_map.get(pgm, ["", "", ""])[0]
        raw3 = self.title_3_map.get(popid, "")

        self.title1_edit.setText(subs(raw1))
        self.title2_edit.setText(subs(raw2))
        self.title3_edit.setText(subs(raw3))

        # Footnotes
        notes = []
        if param_foot: notes.append(subs(param_foot))
        if tbl_foot: notes.append(subs(tbl_foot))
        self.footnote_box.setPlainText("\n".join(notes))

        self.render_header_footer_preview()

        self.grid.setRowCount(3)
        for r in range(3):
            self.grid.setItem(r, 0, QTableWidgetItem(f"Row {r+1}"))
            cb = QComboBox(); cb.addItems(["n", "n(%)", "Mean(SD)", "Min", "Max", "NPCT"])
            self.grid.setCellWidget(r, 1, cb)
            self.grid.setItem(r, 2, QTableWidgetItem("xx"))

    def render_header_footer_preview(self):
        txt = ""
        for sec in ("Header", "Footer"):
            txt += f"--- {sec} ---\n"
            for i in sorted(self.header_footer_data[sec].keys()):
                l, c, r = self.header_footer_data[sec][i]
                txt += f"{l:<30}{c:^40}{r:>30}\n"
            txt += "\n"
        self.header_preview.setPlainText(txt)
        self.footer_preview.setPlainText(txt)

    def add_row(self):
        r = self.grid.rowCount(); self.grid.insertRow(r)
        self.grid.setItem(r, 0, QTableWidgetItem(""))
        cb = QComboBox(); cb.addItems(["n", "n(%)", "Mean(SD)", "Min", "Max", "NPCT"])
        self.grid.setCellWidget(r, 1, cb)
        self.grid.setItem(r, 2, QTableWidgetItem(""))

    def export_rtf(self):
        QMessageBox.information(self, "RTF Export", "RTF export logic not implemented yet.")

    def export_pdf(self):
        rtf_path, _ = QFileDialog.getOpenFileName(self, "Select RTF", "", "RTF Files (*.rtf)")
        if not rtf_path:
            return
        pdf_path, _ = QFileDialog.getSaveFileName(self, "Save PDF", rtf_path.replace(".rtf", ".pdf"), "PDF Files (*.pdf)")
        if not pdf_path:
            return
        try:
            word = win32.gencache.EnsureDispatch("Word.Application")
            doc = word.Documents.Open(rtf_path)
            doc.SaveAs(pdf_path, FileFormat=17)
            doc.Close()
            word.Quit()
            self.status.setText(f"✅ PDF exported: {pdf_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"PDF export failed:\n{e}")

    def export_sas(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save SAS", f"{self.current_table_id}_macro.sas", "SAS Files (*.sas)")
        if not path:
            return
        try:
            code = f"%macro table_{self.current_table_id.replace('.', '_')}();\n"
            for i in range(1, 4):
                title = getattr(self, f"title{i}_edit").toPlainText().strip()
                if title:
                    code += f"title{i} '{title}';\n"
            code += "proc report data=mydata nowd;\ncolumns "
            for r in range(self.grid.rowCount()):
                col = self.grid.item(r, 0)
                if col:
                    text = col.text().strip().replace(" ", "_")
                    code += f"{text} "
            code += ";\nrun;\n%mend;\n"
            with open(path, "w") as f:
                f.write(code)
            self.status.setText(f"✅ SAS macro saved: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"SAS Export failed:\n{e}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = TableShellTool()
    w.show()
    sys.exit(app.exec_())
