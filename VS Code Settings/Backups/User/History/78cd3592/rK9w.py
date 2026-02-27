import sys
import os
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFileDialog,
    QTableWidget, QTableWidgetItem, QGroupBox, QTextEdit, QComboBox, QMessageBox
)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
import win32com.client as win32


class TableShellTool(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("📄 Table Shell Generator")
        self.resize(1200, 900)
        self.excel_data = {}
        self.current_table_id = ""
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()

        # File loader
        file_layout = QHBoxLayout()
        self.file_label = QLabel("No Excel file selected")
        load_btn = QPushButton("Load Excel Spec")
        load_btn.clicked.connect(self.load_excel)
        file_layout.addWidget(self.file_label)
        file_layout.addWidget(load_btn)
        layout.addLayout(file_layout)

        # Table ID dropdown
        tableid_layout = QHBoxLayout()
        self.table_id_dropdown = QComboBox()
        self.table_id_dropdown.currentTextChanged.connect(self.load_table_data)
        tableid_layout.addWidget(QLabel("Select Table ID:"))
        tableid_layout.addWidget(self.table_id_dropdown)
        layout.addLayout(tableid_layout)

        # Minimize buttons
        toggle_layout = QHBoxLayout()
        self.toggle_meta_btn = QPushButton("🔽 Hide Titles + Footnotes")
        self.toggle_header_btn = QPushButton("🔽 Hide Header + Footer")
        self.toggle_meta_btn.clicked.connect(self.toggle_meta_blocks)
        self.toggle_header_btn.clicked.connect(self.toggle_header_footer_blocks)
        toggle_layout.addWidget(self.toggle_meta_btn)
        toggle_layout.addWidget(self.toggle_header_btn)
        layout.addLayout(toggle_layout)

        # Collapsible boxes
        self.header_box = self.make_box("Header")
        self.title_box = self.make_box("Titles")
        self.footnote_box = self.make_box("Footnotes")
        self.footer_box = self.make_box("Footer")

        self.meta_boxes = [self.title_box, self.footnote_box]
        self.hf_boxes = [self.header_box, self.footer_box]

        # Table Content
        self.content_table = QTableWidget()
        self.content_table.setColumnCount(1)
        self.content_table.setHorizontalHeaderLabels(["Table Content"])
        self.content_table.horizontalHeader().setStretchLastSection(True)

        # Add to layout
        layout.addWidget(self.header_box)
        layout.addWidget(self.title_box)
        layout.addWidget(self.content_table)
        layout.addWidget(self.footnote_box)
        layout.addWidget(self.footer_box)

        # Export buttons
        export_layout = QHBoxLayout()
        export_rtf = QPushButton("Export RTF")
        export_rtf.clicked.connect(self.export_rtf)
        export_pdf = QPushButton("Export PDF")
        export_pdf.clicked.connect(self.export_pdf)
        export_layout.addWidget(export_rtf)
        export_layout.addWidget(export_pdf)
        layout.addLayout(export_layout)

        self.setLayout(layout)

    def make_box(self, title):
        box = QGroupBox(title)
        box.setCheckable(True)
        box.setChecked(True)
        layout = QVBoxLayout()
        text_edit = QTextEdit()
        layout.addWidget(text_edit)
        box.setLayout(layout)
        setattr(self, f"{title.lower()}_text", text_edit)
        return box

    def toggle_meta_blocks(self):
        visible = self.meta_boxes[0].isVisible()
        for box in self.meta_boxes:
            box.setVisible(not visible)
        self.toggle_meta_btn.setText("🔼 Show Titles + Footnotes" if visible else "🔽 Hide Titles + Footnotes")

    def toggle_header_footer_blocks(self):
        visible = self.hf_boxes[0].isVisible()
        for box in self.hf_boxes:
            box.setVisible(not visible)
        self.toggle_header_btn.setText("🔼 Show Header + Footer" if visible else "🔽 Hide Header + Footer")

    def load_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Open Excel Spec", "", "Excel Files (*.xlsx)")
        if not path:
            return
        self.file_label.setText(os.path.basename(path))
        wb = load_workbook(path, data_only=True)

        try:
            self.excel_data = {
                "Tables": {row[3].value: row for row in wb["Tables"].iter_rows(min_row=2) if row[3].value},
                "Table Runs": [r for r in wb["Table Runs"].iter_rows(min_row=2)],
                "Populations": {str(r[0].value): r[2].value for r in wb["Populations"].iter_rows(min_row=2) if r[0].value},
                "Parameters": {str(r[0].value): r[3].value for r in wb["Parameters"].iter_rows(min_row=2) if r[0].value},
                "HeaderFooter": self.parse_hf(wb)
            }
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load Excel: {e}")
            return

        self.table_id_dropdown.clear()
        ids = sorted({str(r[1].value).strip() for r in self.excel_data["Table Runs"] if r[1].value})
        self.table_id_dropdown.addItems(ids)

    def parse_hf(self, wb):
        for name in ["Headersfooters", "Headers and Footers"]:
            if name in wb:
                d = {"Header": [], "Footer": []}
                for r in wb[name].iter_rows(min_row=2):
                    kind = str(r[0].value).strip().capitalize()
                    try:
                        ln = int(r[1].value)
                    except:
                        continue
                    txt = f"{r[2].value or '':<40}{r[3].value or '':^40}{r[4].value or '':>40}".rstrip()
                    if kind in d:
                        d[kind].append((ln, txt))
                return d
        return {"Header": [], "Footer": []}

    def load_table_data(self, table_id):
        self.current_table_id = table_id
        run = next((r for r in self.excel_data["Table Runs"] if str(r[1].value).strip() == table_id), None)
        if not run:
            return
        pgm = run[3].value
        popid = str(run[4].value or "").strip()
        paramid = str(run[5].value or "").strip()
        param_label = self.excel_data["Parameters"].get(paramid, "")

        table_row = self.excel_data["Tables"].get(pgm, [])
        # Titles block: Line 1, 2, 3
        title1 = f"Table {table_id}"
        title2_parts = [table_row[9].value, table_row[10].value, table_row[11].value] if table_row else []
        title2 = " ".join([t for t in title2_parts if t]).replace("&ParameterLabel", param_label)
        title3 = self.excel_data["Populations"].get(popid, "")
        self.titles_text.setPlainText("\n".join([title1, title2, title3]))

        # Footnotes
        footnotes = [table_row[i].value or "" for i in range(12, 15)] if len(table_row) >= 15 else []
        footnotes = [f.replace("&ParameterLabel", param_label) for f in footnotes]
        self.footnotes_text.setPlainText("\n".join([f for f in footnotes if f]))

        # Header/Footer
        header = [t for _, t in sorted(self.excel_data["HeaderFooter"].get("Header", []))]
        footer = [t for _, t in sorted(self.excel_data["HeaderFooter"].get("Footer", []))]
        self.header_text.setPlainText("\n".join(header))
        self.footer_text.setPlainText("\n".join(footer))

        # Dummy Table Content
        self.content_table.setRowCount(3)
        self.content_table.setItem(0, 0, QTableWidgetItem("-" * 100))
        self.content_table.setItem(1, 0, QTableWidgetItem("Sample table content"))
        self.content_table.setItem(2, 0, QTableWidgetItem("-" * 100))

    def export_rtf(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save RTF", f"{self.current_table_id}.rtf", "RTF Files (*.rtf)")
        if not path:
            return
        try:
            with open(path, "w") as f:
                f.write("{\\rtf1\\ansi\\landscape\\deff0{\\fonttbl{\\f0 Courier New;}}\n")
                f.write("\\fs16\n")
                for section in [
                    self.header_text, self.titles_text, self.content_table, self.footnotes_text, self.footer_text
                ]:
                    if isinstance(section, QTextEdit):
                        for line in section.toPlainText().splitlines():
                            f.write(f"{line}\\line\n")
                    elif isinstance(section, QTableWidget):
                        for i in range(section.rowCount()):
                            item = section.item(i, 0)
                            if item:
                                f.write(f"{item.text()}\\line\n")
                f.write("}")
            QMessageBox.information(self, "RTF Saved", f"Saved: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def export_pdf(self):
        pdf_path, _ = QFileDialog.getSaveFileName(self, "Save PDF", f"{self.current_table_id}.pdf", "PDF Files (*.pdf)")
        if not pdf_path:
            return
        rtf_temp = pdf_path.replace(".pdf", ".rtf")
        self.export_rtf_internal(rtf_temp)
        try:
            word = win32.gencache.EnsureDispatch('Word.Application')
            doc = word.Documents.Open(rtf_temp)
            doc.SaveAs(pdf_path, FileFormat=17)
            doc.Close()
            word.Quit()
            os.remove(rtf_temp)
            QMessageBox.information(self, "PDF Saved", f"Saved: {pdf_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def export_rtf_internal(self, path):
        try:
            with open(path, "w") as f:
                f.write("{\\rtf1\\ansi\\landscape\\deff0{\\fonttbl{\\f0 Courier New;}}\n")
                f.write("\\fs16\n")
                for section in [
                    self.header_text, self.titles_text, self.content_table, self.footnotes_text, self.footer_text
                ]:
                    if isinstance(section, QTextEdit):
                        for line in section.toPlainText().splitlines():
                            f.write(f"{line}\\line\n")
                    elif isinstance(section, QTableWidget):
                        for i in range(section.rowCount()):
                            item = section.item(i, 0)
                            if item:
                                f.write(f"{item.text()}\\line\n")
                f.write("}")
        except:
            pass


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = TableShellTool()
    window.show()
    sys.exit(app.exec_())
