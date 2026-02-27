import sys, os, re
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QComboBox, QTextEdit, QFileDialog, QLineEdit, QMessageBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QFrame
)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
import win32com.client as win32

class TableShellTool(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("📄 Table Shell Exporter")
        self.resize(1200, 900)
        self.excel_data = {}
        self.param_labels = {}
        self.param_foot = {}
        self.title_2_map = {}
        self.table_foot = {}
        self.title_3_map = {}
        self.current_table_id = ""
        self.header_footer_data = {"Header": {}, "Footer": {}}
        self.header_footer_visible = True
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()
        layout.addWidget(QLabel("<h2>📘 Clinical Table Shell Builder</h2>"))

        # File selection
        file_row = QHBoxLayout()
        self.excel_path = QLineEdit(); self.excel_path.setReadOnly(True)
        browse_btn = QPushButton("Browse Excel Spec"); browse_btn.clicked.connect(self.load_excel)
        file_row.addWidget(self.excel_path); file_row.addWidget(browse_btn)
        layout.addLayout(file_row)

        # Table selector
        layout.addWidget(QLabel("Select Table ID:"))
        self.table_dropdown = QComboBox()
        self.table_dropdown.currentTextChanged.connect(self.load_table_content)
        layout.addWidget(self.table_dropdown)

        # Toggle header/footer
        self.toggle_header_footer_btn = QPushButton("🔽 Show/Hide Header/Footer")
        self.toggle_header_footer_btn.clicked.connect(self.toggle_header_footer)
        layout.addWidget(self.toggle_header_footer_btn)

        # Header/Footer preview
        self.header_footer_container = QFrame()
        hf_layout = QVBoxLayout()
        self.header_preview = QTextEdit(); self.header_preview.setReadOnly(True)
        self.footer_preview = QTextEdit(); self.footer_preview.setReadOnly(True)
        hf_layout.addWidget(QLabel("Header Preview:")); hf_layout.addWidget(self.header_preview)
        hf_layout.addWidget(QLabel("Footer Preview:")); hf_layout.addWidget(self.footer_preview)
        self.header_footer_container.setLayout(hf_layout)
        layout.addWidget(self.header_footer_container)

        # Title lines
        for i, label in enumerate(["Title Line 1", "Title Line 2", "Title Line 3"]):
            layout.addWidget(QLabel(label))
            te = QTextEdit(); te.setFixedHeight(30)
            setattr(self, f"title{i+1}_edit", te)
            layout.addWidget(te)

        # Footnote box
        layout.addWidget(QLabel("Footnotes:"))
        self.footnote_box = QTextEdit(); self.footnote_box.setReadOnly(True); self.footnote_box.setFixedHeight(60)
        layout.addWidget(self.footnote_box)

        # Shell grid
        self.grid = QTableWidget(0, 3)
        self.grid.setHorizontalHeaderLabels(["Shell Row Text", "Format", "Dummy Value"])
        self.grid.setEditTriggers(QAbstractItemView.AllEditTriggers)
        self.grid.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.grid)

        add_btn = QPushButton("➕ Add Row"); add_btn.clicked.connect(self.add_row)
        layout.addWidget(add_btn)

        # Export buttons
        export_row = QHBoxLayout()
        for label, method in [("📄 Export RTF", self.export_rtf),
                              ("📄 Export PDF", self.export_pdf),
                              ("💻 Export SAS Macro", self.export_sas)]:
            btn = QPushButton(label); btn.clicked.connect(method); export_row.addWidget(btn)
        layout.addLayout(export_row)

        # Status
        self.status = QLabel("Ready."); layout.addWidget(self.status)
        self.setLayout(layout)

    def toggle_header_footer(self):
        self.header_footer_visible = not self.header_footer_visible
        self.header_footer_container.setVisible(self.header_footer_visible)

    def load_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel Spec", "", "Excel Files (*.xlsx *.xls)")
        if not path:
            return
        self.excel_path.setText(path)
        wb = load_workbook(path, data_only=True)
        sheets = {s.lower().replace(" ", ""): s for s in wb.sheetnames}
        get = lambda name: sheets.get(name.lower().replace(" ", ""))

        runs = wb[get("TableRuns")]
        tables = wb[get("Tables")]
        pops = wb[get("Populations")]
        params = wb[get("Parameters")]
        header_sheet = get("HeadersFooters") or get("HeadersandFooters")
        headers = wb[header_sheet] if header_sheet else None

        # Clear previous data
        for d in [self.excel_data, self.param_labels, self.param_foot,
                  self.title_2_map, self.table_foot, self.title_3_map]:
            d.clear()
        self.table_dropdown.clear()
        self.header_footer_data = {"Header": {}, "Footer": {}}

        # Load runs sheet
        for row in runs.iter_rows(min_row=2, values_only=True):
            tid, pgm, popid, pid = row[1], row[4], row[6], row[9]
            if tid and pgm:
                tid, pgm, popid, pid = str(tid).strip(), str(pgm).strip(), str(popid or "").strip(), str(pid or "").strip()
                self.excel_data[tid] = {"pgm": pgm, "popid": popid, "pid": pid}
                self.table_dropdown.addItem(tid)

        # Load tables sheet
        for row in tables.iter_rows(min_row=2, values_only=True):
            pgm = str(row[3]).strip() if row[3] else None
            title = str(row[9] or "")
            foot = str(row[10] or "")
            if pgm:
                self.title_2_map[pgm] = title
                self.table_foot[pgm] = foot

        # Load pops
        for row in pops.iter_rows(min_row=2, values_only=True):
            pid, name = str(row[0]).strip(), str(row[2] or "")
            self.title_3_map[pid] = name

        # Load params
        for row in params.iter_rows(min_row=2, values_only=True):
            pid = str(row[0]).strip()
            label = str(row[3] or "")
            foot = str(row[4] or "")  # assuming Paramvar_fn in col5
            self.param_labels[pid] = label
            self.param_foot[pid] = foot

        # Headers
        if headers:
            for r in headers.iter_rows(min_row=2, values_only=True):
                try:
                    section, line = r[0], int(r[1])
                    if section in ["Header","Footer"]:
                        self.header_footer_data[section].setdefault(line, ["","",""])
                        for i in range(3):
                            if r[i+2]:
                                self.header_footer_data[section][line][i] = str(r[i+2]).strip()
                except:
                    continue

    def load_table_content(self, tid):
        if not tid:
            return
        self.current_table_id = tid
        self.grid.setRowCount(0)

        info = self.excel_data[tid]
        pgm, popid, pid = info["pgm"], info["popid"], info["pid"]
        param_label = self.param_labels.get(pid, "")

        def subs(txt):
            return re.sub(r"&ParameterLabel", param_label, txt or "")

        title1 = subs(f"Table {tid}")
        title2 = subs(self.title_2_map.get(pgm,""))
        title3 = subs(self.title_3_map.get(popid,""))

        self.title1_edit.setText(title1)
        self.title2_edit.setText(title2)
        self.title3_edit.setText(title3)

        # Footnotes
        footnotes = []
        if pid and self.param_foot.get(pid):
            footnotes.append(subs(self.param_foot[pid]))
        if self.table_foot.get(pgm):
            footnotes.append(subs(self.table_foot[pgm]))
        self.footnote_box.setPlainText("\n".join(footnotes))

        # Header/Footer preview
        self.render_header_footer_preview()
        self.header_footer_container.setVisible(self.header_footer_visible)

        # Initialize grid
        self.grid.setRowCount(3)
        for r in range(3):
            self.grid.setItem(r, 0, QTableWidgetItem(f"Row {r+1}"))
            fmt = QComboBox(); fmt.addItems(["n","n(%)","Mean(SD)","Min","Max","NPCT"])
            self.grid.setCellWidget(r, 1, fmt)
            self.grid.setItem(r, 2, QTableWidgetItem("xx"))

    def render_header_footer_preview(self):
        def join_lines(data):
            lines=[]
            for i in sorted(data.keys()):
                l,c,r = data[i]
                lines.append(f"{l:<30}{c:^40}{r:>30}")
            return "\n".join(lines)
        self.header_preview.setText(join_lines(self.header_footer_data["Header"]))
        self.footer_preview.setText(join_lines(self.header_footer_data["Footer"]))

    def add_row(self):
        r = self.grid.rowCount(); self.grid.insertRow(r)
        self.grid.setItem(r,0,QTableWidgetItem("")); fmt=QComboBox()
        fmt.addItems(["n","n(%)","Mean(SD)","Min","Max","NPCT"])
        self.grid.setCellWidget(r,1,fmt); self.grid.setItem(r,2,QTableWidgetItem("xx"))

    def export_rtf(self):
        # implementation omitted for brevity
        QMessageBox.information(self, "Export", "RTF export not implemented in this snippet.")

    def export_pdf(self):
        QMessageBox.information(self, "Export", "PDF export not implemented in this snippet.")

    def export_sas(self):
        QMessageBox.information(self, "Export", "SAS export not implemented in this snippet.")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = TableShellTool(); w.show()
    sys.exit(app.exec_())
