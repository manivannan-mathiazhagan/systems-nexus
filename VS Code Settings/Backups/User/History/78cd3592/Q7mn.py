import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
import win32com.client as win32

class TableShellTool(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("🧾 Table Shell Exporter v4.4")
        self.resize(1000, 700)
        self.excel_data = {}
        self.title2_map = {}
        self.title3_map = {}
        self.param_labels = {}
        self.header_footer_data = {"Header": {}, "Footer": {}}
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()
        layout.addWidget(QLabel("<h2>Clinical Table Shell Builder</h2>"))

        file_row = QHBoxLayout()
        self.excel_path = QLineEdit()
        self.excel_path.setReadOnly(True)
        btn = QPushButton("Browse Excel Spec")
        btn.clicked.connect(self.load_excel)
        file_row.addWidget(self.excel_path); file_row.addWidget(btn)
        layout.addLayout(file_row)

        layout.addWidget(QLabel("Select Table ID:"))
        self.table_dropdown = QComboBox()
        self.table_dropdown.currentTextChanged.connect(self.load_table_content)
        layout.addWidget(self.table_dropdown)

        hfc_btn = QPushButton("🔽 Toggle Header/Footer Preview")
        hfc_btn.clicked.connect(lambda: self.header_footer_container.setVisible(
            not self.header_footer_container.isVisible()))
        layout.addWidget(hfc_btn)

        self.header_footer_container = QFrame()
        hfl = QVBoxLayout()
        self.header_preview = QTextEdit(); self.footer_preview = QTextEdit()
        self.header_preview.setReadOnly(True); self.footer_preview.setReadOnly(True)
        hfl.addWidget(QLabel("Header Preview:")); hfl.addWidget(self.header_preview)
        hfl.addWidget(QLabel("Footer Preview:")); hfl.addWidget(self.footer_preview)
        self.header_footer_container.setLayout(hfl)
        layout.addWidget(self.header_footer_container)

        for i in range(1, 4):
            lbl = QLabel(f"Title Line {i}")
            te = QTextEdit(); te.setFixedHeight(30)
            setattr(self, f"title{i}_edit", te)
            layout.addWidget(lbl); layout.addWidget(te)

        self.grid = QTableWidget(0, 3)
        self.grid.setHorizontalHeaderLabels(["Text", "Format", "Value"])
        self.grid.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.grid)

        add_btn = QPushButton("➕ Add Row")
        add_btn.clicked.connect(self.add_row)
        layout.addWidget(add_btn)

        erow = QHBoxLayout()
        for text, method in [("Export RTF", self.export_rtf),
                             ("Export PDF", self.export_pdf),
                             ("Export SAS Macro", self.export_sas)]:
            b = QPushButton(text)
            b.clicked.connect(method)
            erow.addWidget(b)
        layout.addLayout(erow)

        self.status = QLabel("Ready")
        layout.addWidget(self.status)
        self.setLayout(layout)

    def add_row(self):
        row = self.grid.rowCount()
        self.grid.insertRow(row)
        self.grid.setItem(row, 0, QTableWidgetItem(""))
        cb = QComboBox()
        cb.addItems(["n", "n(%)", "Mean(SD)", "Min", "Max", "NPCT"])
        self.grid.setCellWidget(row, 1, cb)
        self.grid.setItem(row, 2, QTableWidgetItem("xx"))

    def load_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel Spec", "", "Excel files (*.xlsx *.xls)")
        if not path:
            return
        self.excel_path.setText(path)
        wb = load_workbook(path, data_only=True)
        s = {sh.lower().replace(" ", ""): sh for sh in wb.sheetnames}
        rt = wb[s["tableruns"]]
        tb = wb[s["tables"]]
        pop = wb[s["populations"]]
        hf = wb[s.get("headersfooters") or s.get("headersandfooters")]
        prm = wb[s["parameters"]]

        self.excel_data.clear(); self.param_labels.clear()
        self.title2_map.clear(); self.title3_map.clear()
        self.header_footer_data = {"Header": {}, "Footer": {}}

        for r in rt.iter_rows(min_row=2, values_only=True):
            tid = str(r[1]).strip()
            if not tid:
                continue
            paramids = str(r[2] or "").strip()
            pgm = str(r[4] or "").strip()
            popid = str(r[6] or "").strip()
            self.excel_data[tid] = {
                "paramids": paramids.split(),
                "pgm": pgm,
                "popid": popid
            }
            self.table_dropdown.addItem(tid)

        for r in tb.iter_rows(min_row=2, values_only=True):
            pgm = str(r[3]).strip()
            if pgm:
                titles = [str(x).strip() for x in (r[9], r[10], r[11]) if x]
                self.title2_map[pgm] = " - ".join(titles)

        for r in pop.iter_rows(min_row=2, values_only=True):
            pid = str(r[0]).strip()
            lbl = str(r[2]).strip()
            if pid:
                self.title3_map[pid] = lbl

        for r in prm.iter_rows(min_row=2, values_only=True):
            pid = str(r[0]).strip()
            lbl = str(r[1]).strip()
            if pid:
                self.param_labels[pid] = lbl

        for r in hf.iter_rows(min_row=2, values_only=True):
            sect = str(r[0]).strip()
            line = int(r[1])
            left = str(r[2] or "")
            center = str(r[3] or "")
            right = str(r[4] or "")
            if sect in ("Header", "Footer"):
                self.header_footer_data[sect][line] = (left, center, right)

        self.status.setText("Excel loaded successfully")

    def load_table_content(self, tid):
        self.grid.setRowCount(0)
        data = self.excel_data.get(tid, {})
        paramids = data.get("paramids", [])
        pgm = data.get("pgm", "")
        popid = data.get("popid", "")

        def sub(t): return t.replace("&ParameterLabel",
            ", ".join(self.param_labels.get(pid, "") for pid in paramids))

        self.title1_edit.setText(sub(f"Table {tid}"))
        self.title2_edit.setText(sub(self.title2_map.get(pgm, "")))
        self.title3_edit.setText(sub(self.title3_map.get(popid, "")))

        # render previews
        def fmt_preview(d):
            lines = []
            for ln in sorted(d):
                l,c,r = d[ln]
                lines.append(f"{l:<30}{c:^40}{r:>30}".rstrip())
            return "\n".join(lines)

        self.header_preview.setText(fmt_preview(self.header_footer_data["Header"]))
        self.footer_preview.setText(fmt_preview(self.header_footer_data["Footer"]))

        for i in range(3):
            self.add_row()

        self.status.setText(f"Loaded table {tid}")

    def export_rtf(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save RTF File", f"{self.table_dropdown.currentText()}.rtf", "RTF Files (*.rtf)")
        if not path:
            return
        tid = self.table_dropdown.currentText()
        data = self.excel_data.get(tid, {})
        paramids = data.get("paramids", [])

        header_txt = "".join(
            rf"\pard\ql\f0\fs16 {l.strip()} \tab\qc {c.strip()} \tab\qr {r.strip()}\par"
            for l,c,r in self.header_footer_data["Header"].values()
        )
        footer_txt = "".join(
            rf"\pard\ql\f0\fs16 {l.strip()} \tab\qc {c.strip()} \tab\qr {r.strip()}\par"
            for l,c,r in self.header_footer_data["Footer"].values()
        )

        lines = [r"{\rtf1\ansi\deff0", r"{\fonttbl{\f0 Courier New;}}", r"\fs16",
                 r"\paperw16840\paperh11900\landscape", r"\margl1440\margr1440\margt1440\margb1440",
                 "{\\header " + header_txt + "}", "{\\footer " + footer_txt + "}"]

        for i in range(1,4):
            txt = getattr(self, f"title{i}_edit").toPlainText().strip()
            if txt:
                txt = txt.replace("&ParameterLabel",
                                   ", ".join(self.param_labels.get(pid,"") for pid in paramids))
                lines.append(rf"\pard\qc {txt}\par")

        lines.append(r"\trowd\cellx3000\cellx6000\cellx9000")
        for r in range(self.grid.rowCount()):
            t = self.grid.item(r,0).text()
            f = self.grid.cellWidget(r,1).currentText()
            v = self.grid.item(r,2).text()
            lines.append(rf"\intbl\f0\fs16 {t}\cell {f}\cell {v}\cell\row")
        lines.append("}")

        with open(path, "w") as f:
            f.write("\n".join(lines))

        self.status.setText(f"✅ RTF saved: {path}")

    def export_pdf(self):
        rtf, _ = QFileDialog.getOpenFileName(self, "Select RTF to Convert", "", "RTF Files (*.rtf)")
        if not rtf:
            return
        pdf, _ = QFileDialog.getSaveFileName(self, "Save PDF File", rtf.replace(".rtf", ".pdf"), "PDF Files (*.pdf)")
        if not pdf:
            return
        try:
            wd = win32.gencache.EnsureDispatch("Word.Application")
            doc = wd.Documents.Open(rtf)
            doc.SaveAs(pdf, FileFormat=17)
            doc.Close(); wd.Quit()
            self.status.setText(f"✅ PDF exported: {pdf}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"PDF export failed:\n{e}")

    def export_sas(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save SAS Macro", f"{self.table_dropdown.currentText()}.sas", "SAS Files (*.sas)")
        if not path:
            return
        try:
            code = f"%macro table_{self.table_dropdown.currentText().replace('.','_')}();\n"
            for i in range(1,4):
                t = getattr(self,f"title{i}_edit").toPlainText().strip()
                if t:
                    code += f"title{i} '{t}';\n"
            code += "proc report data=mydata nowd;\ncolumns "
            for r in range(self.grid.rowCount()):
                t = self.grid.item(r,0).text()
                code += f"{t.replace(' ','_')} "
            code += ";\nrun;\n%mend;\n"
            with open(path, "w") as f:
                f.write(code)
            self.status.setText(f"✅ SAS saved: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"SAS export failed:\n{e}")

if __name__=="__main__":
    app = QApplication(sys.argv)
    w = TableShellTool()
    w.show()
    sys.exit(app.exec_())
